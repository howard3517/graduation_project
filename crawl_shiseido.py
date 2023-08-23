from time import sleep
#from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import requests 
from bs4 import BeautifulSoup as bs
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os 


header = {'User-Agen':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15'}
url = 'https://www.global-shiseido.com.tw'

def crawler(url,dir,producttype,filepath,writer):
    try:
        print(f'{producttype} clawing ... ')
        try:
            req = requests.get(url+ dir +'?srule=most-popular&sz=156',headers=header)
            data = bs(req.text,'html.parser')
        except:
            try:
                req = requests.get(url+ dir +'?srule=most-popular&sz=36',headers=header)
                data = bs(req.text,'html.parser')
            except:
                req = requests.get(url+ dir +'?srule=most-popular&sz=24',headers=header)
                data = bs(req.text,'html.parser')
        
        productname = []
        price = []
        pic = []
        brand = []
        info_href = []
        info_content = []
        check = [] #所有抓到產品內文的row
        amount = 0 #總共有幾個產品

        # get product name
        for product in data.select('h5.product-name'):
            productname.append(''.join(product.text.split('\n')))
            
        # get product price
        for product in data.find_all('h2'):
            temp_text = ''.join(product.text.split('\n'))
            temp_text = ''.join(temp_text.split(','))
            price.append(''.join(temp_text.split('NT$')))
        
        # get product info href
        for product in data.select('a.thumb-link'):
            info_href.append(product.get('href'))
        #get product info content web code(html) 
        for count,content in enumerate(info_href):
            link = requests.get(url+str(content),headers=header)
            info = bs(link.text,'html.parser')
            # get info content   
            for content in info.select(' div.product-info > div.product-description.mobile-only > span'):# 程式碼不一致導致會有些路徑不是這樣
                check.append(count)
                info_content.append(content.text)
            if content == []: #
                for content in info.select('div.product-description.desktop-only > span > p'):
                    check.append(count)
                    info_content.append(content.text)

        print(f'content amount : {len(check)}')


        # get product pic
        for product in data.select('img.thumb-image'):
            pic.append(product.get('src'))
            
        #get product brand
        for product in data.select('h4.product-brand'):
            brand.append(''.join(product.string.split('\n')))
            
        # make table
        df = pd.DataFrame()
        df['name'] = productname
        df['price'] = price
        df['brand'] = brand
        df['pic'] = pic
        df['content'] = info_content
        
        
        # save data to excel
        df.to_excel(writer, sheet_name = producttype ,index=False)

        
        
        #check amount of product
        crawlamount= df.shape[0]
        temp = data.select('#results-hits-top > span')
        temp = str(temp)
        temp = ''.join(temp.split('<span>'))
        temp = ''.join(temp.split('</span>'))
        temp = ''.join(temp.split('['))
        temp = ''.join(temp.split(']'))
        shouldamount = int(temp)
        
        if crawlamount == shouldamount:
            print(f'{producttype} : Done \nTotal amount : {shouldamount}\n')
        else:
            print(f'{producttype} : got wrong \nactual amount : {shouldamount}\ncrawl amount : {crawlamount}\n')
    except Exception as e:
        print(e)
        
        
        
        
def get_type(url,filepath,writer):
    res = requests.get(url)
    text = bs(res.text,'html.parser')
    
    producttype = []
    href = []
    #original : navigation > div > ul > li:nth-child(4) > div > div.level-2-wrapper > div.level-2-full-width > div > ul > li:nth-child(4) > ul > li > ul > li:nth-child(3) > a
    for name in text.select(' div > div.level-2-wrapper > div.level-2-full-width > div > ul > li > ul > li > ul > li > a'):
        temp_text = ' '.join(name.text.split('/')) # excel sheet name can't include /
        producttype.append(''.join(temp_text.split('\n'))) # 去除品名內的換行符號
        href.append(name.get('href'))
    
    df = pd.DataFrame()
    df['type'] = producttype
    df['href'] = href
    
    df.drop(index=0,inplace=True)#第一個頁面是空的不用爬 
    #刪除後 index 從 1 開始
    

    
    df.to_excel(writer, sheet_name='Summary' ,index=False)
    
    print(f'Summary finish!\n')
    
    return df

def crawl_rest(df):
    print('第二度開始爬蟲')
    os.mkdir('temp')

    # 去掉含有 all 的項目
    all_list = df[df['type'].str.contains('All')==False]['type'].tolist()
    # 抓爬到的產品
    file = openpyxl.load_workbook('Shiseido.xlsx')
    got_list = file.sheetnames

    # 沒爬到的
    rest = list(set(all_list) - set(got_list))

    for series in rest:
        try:
            row = df[df['type']==series].index[0]
            link = df['href'][row]
            data = requests.get(url+link+'?srule=most-popular&sz=156',headers=header)
            data = bs(data.text,'html.parser')
            #print(data)

            # 將所有子產品的url存起來
            info_href = []
            for product in data.select('a.thumb-link'):
                info_href.append(product.get('href'))
                #print(product.get('href'))
            print(f'{series} total:{len(info_href)}')

            # 到各子頁面抓子產品內容
            comment = []
            for path in info_href:  
                data = requests.get(url+link+path,headers=header)
                data = bs(data.text,'html.parser')
                #data


                for name in data.select(' div.product-info > div.product-description.mobile-only > span'):
                    comment.append(name.text)
                if name == []:
                    pass
                else:
                    for name in data.select(' div.product-description.desktop-only > span > p > a '):
                        comment.append(name.text)
            print(f'comment amount:{len(comment)}')

            table = pd.read_excel('Shiseido_所有資產品的名稱.xlsx',series)
            table['content'] = comment
            table.to_excel('temp/'+series+'.xlsx',index=False)
        except Exception as e:
            print(e)

    sec_time = []
    for file in os.listdir('temp'):
        sec_time.append(file[:-5])

    return list(set(rest) - set(sec_time))

def combine(name):
    writer = pd.ExcelWriter(name, engine = 'openpyxl')
    for file in os.listdir('temp'):
        data = pd.read_excel('temp/'+file)
        #print(data)
        data.to_excel(writer,sheet_name = file[:-5],index=False)

    writer.close()    

if __name__ == '__main__':
    
    #create a new excel to save crawler's data
    workbook = openpyxl.Workbook()
    workbook.save('Shiseido.xlsx')
    
    #open excel file
    filepath = 'Shiseido.xlsx'
    book = load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    
    summary = get_type(url,'Shiseido.xlsx',writer)

    
    for i in range(1,summary.shape[0]):
        # index 11 新艷陽．夏 架構不一樣
        if i != 11:
            crawler(url,summary['href'][i],summary['type'][i],'Shiseido.xlsx',writer)
        
    #excel 要一直開著才能在同個檔案保留舊的並新增新的 sheet
    writer.close()


    # 最後剩下的產品
    final = rest_product = crawl_rest(summary)
    print(f'最終剩下沒爬：\n{final}')

    combine('Shiseido_2.xlsx')


    print('Finish !!!')
    
    

        
    

