import requests 
from bs4 import BeautifulSoup as bs
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os 

header = {'User-Agen':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15'}
url = 'https://www.global-shiseido.com.tw'


def get_summary(url,writer):
    res = requests.get(url)
    text = bs(res.text,'html.parser')
    
    producttype = []
    href = []
    #original : navigation > div > ul > li:nth-child(4) > div > div.level-2-wrapper > div.level-2-full-width > div > ul > li:nth-child(4) > ul > li > ul > li:nth-child(3) > a
    for name in text.select(' div > div.level-2-wrapper > div.level-2-full-width > div > ul > li > ul > li > ul > li > a'):
        temp_text = ' '.join(name.text.split('/')) # excel sheet name can't include /
        producttype.append(''.join(temp_text.split('\n'))) # 去除品名內的換行符號
        href.append(name.get('href'))
    
    df = pd.DataFrame()
    df['type'] = producttype
    df['href'] = href
    
    df.drop(index=0,inplace=True)#第一個頁面是空的不用爬 
    #刪除後 index 從 1 開始
    

    
    df.to_excel(writer, sheet_name='Summary' ,index=False)
    
    print(f'Summary done!\n')
    
    return df


def crawl_without_comment(url,dir,producttype,writer):
    print(f'{producttype} clawing ... ')
    try:
        req = requests.get(url+ dir +'?srule=most-popular&sz=156',headers=header)
        data = bs(req.text,'html.parser')
        
        productname = []
        price = []
        pic = []
        brand = []

        # get product name
        for product in data.select('h5.product-name'):
            productname.append(''.join(product.text.split('\n')))
            
        # get product price
        for product in data.find_all('h2'):
            temp_text = ''.join(product.text.split('\n'))
            temp_text = ''.join(temp_text.split(','))
            price.append(''.join(temp_text.split('NT$')))
        
        # get product pic
        for product in data.select('img.thumb-image'):
            pic.append(product.get('src'))
            
        #get product brand
        for product in data.select('h4.product-brand'):
            brand.append(''.join(product.string.split('\n')))
            
        # make table
        df = pd.DataFrame()
        df['name'] = productname
        df['price'] = price
        df['brand'] = brand
        df['pic'] = pic
        
        
        # save data to excel
        df.to_excel(writer, sheet_name = producttype ,index=False)

        
        
        #check amount of product
        crawlamount= df.shape[0]
        temp = data.select('#results-hits-top > span')
        temp = str(temp)
        temp = ''.join(temp.split('<span>'))
        temp = ''.join(temp.split('</span>'))
        temp = ''.join(temp.split('['))
        temp = ''.join(temp.split(']'))
        shouldamount = int(temp)
        
        if crawlamount == shouldamount:
            print(f'OK \nTotal amount : {shouldamount}\n')
        else:
            print(f'got wrong \nactual amount : {shouldamount}\ncrawl amount : {crawlamount}\n')
    except Exception as e:
        print(e)



def crawl_comment(df,file_path):
    print('開始爬內文')

    # 去掉含有 all 的項目
    all_list = df[df['type'].str.contains('All')==False]['type'].tolist()

    for series in all_list:
        try:
            row = df[df['type']==series].index[0]
            link = df['href'][row]
            data = requests.get(url+link+'?srule=most-popular&sz=156',headers=header)
            data = bs(data.text,'html.parser')
            #print(data)

            # 將所有子產品的url存起來
            info_href = []
            for product in data.select('a.thumb-link'):
                info_href.append(product.get('href'))
                #print(product.get('href'))
            print(f'{series} total:{len(info_href)}')

            # 到各子頁面抓子產品內容
            comment = ['content'] # column name
            for path in info_href:  
                data = requests.get(url+link+path,headers=header)
                data = bs(data.text,'html.parser')
                #data


                for name in data.select(' div.product-info > div.product-description.mobile-only > span'):
                    comment.append(name.text)
                if name == []:
                    pass
                else:
                    for name in data.select(' div.product-description.desktop-only > span > p > a '):
                        comment.append(name.text)
            print(f'comment amount:{len(comment)}')

            workbook = load_workbook(file_path)
            worksheet = workbook[series]
            for index,text in enumerate(comment):
                            # index 從0開始 所以要+1
                worksheet.cell(row=index+1,column=6).value = text

            if len(info_href) == len(comment)-1:
                print('OK\n')
            else:
                print('got wrong\n')

        except Exception as e:
            print(e)

    return all_list


def final_check(all,file_path):
    print('最終沒爬到的產品：')
    file = openpyxl.load_workbook(file_path)
    got = file.sheetnames
    ls = list(set(all)-set(got))

    print(ls)


if __name__ == '__main__':

    filepath = 'Shiseido.xlsx'

    #create a new excel to save crawler's data
    workbook = openpyxl.Workbook()
    workbook.save(filepath)
    
    #open excel file
    
    book = load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    
    summary = get_summary(url,writer)

    
    for i in range(1,summary.shape[0]):
        crawl_without_comment(url,summary['href'][i],summary['type'][i],writer)
    print("第一次爬蟲完成")

    writer.close()

    all_list = crawl_comment(summary,filepath)

    final_check(all_list,filepath)

    #excel 要一直開著才能在同個檔案保留舊的並新增新的 sheet
    

    print('Finish !!!!!')